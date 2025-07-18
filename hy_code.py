import os
import shutil
import sys
import pdfplumber
import pandas as pd
import win32print
import win32ui
import win32con
import re
import time
import threading
import queue
import portalocker
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Text, Scrollbar
from reportlab.pdfgen import canvas
from pathlib import Path
import concurrent

# 设置 BASE_DIR 为 .exe 或脚本所在目录
BASE_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))

# 设置文件路径
LOG_FILE = os.path.join(BASE_DIR, "log.txt")
DATA_LOG_FILE = os.path.join(BASE_DIR, "data_log.xlsx")
DESC_FILE = os.path.join(BASE_DIR, "desc.xlsx")
DEFAULT_TEMPLATE_PATH = os.path.join(BASE_DIR, "updated_template.xlsx")

# 全局变量
PDF_SOURCE_DIR = os.path.join(BASE_DIR, "Cirris_Reports")
last_updated_template = DEFAULT_TEMPLATE_PATH
log_text = None
root = None
auto_button = None
use_onenote_var = None
path_label = None
label_count_var = None
result_queue = queue.Queue()

def setup_test_environment():
    test_base_dir = os.path.join(BASE_DIR, "test_Cirris_Reports")
    test_pdf_dir = os.path.join(test_base_dir, "subfolder")
    sample_pdf = os.path.join(test_pdf_dir, "LBTestSample.pdf")

    try:
        test_file = os.path.join(BASE_DIR, "test_write.txt")
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
    except Exception as e:
        log_message(f"BASE_DIR 无写权限: {str(e)}")
        raise PermissionError(f"无法在 {BASE_DIR} 创建测试文件，请检查权限")

    os.makedirs(test_pdf_dir, exist_ok=True)

    try:
        c = canvas.Canvas(sample_pdf)
        c.drawString(100, 750, "NO.: 123")
        c.drawString(100, 735, "Test Name: TestSample")
        c.drawString(100, 720, "Series number: SN12345")
        c.drawString(100, 705, "Lot ID: LOT001")
        c.drawString(100, 690, "Test Date: 2025-05-19")
        c.drawString(100, 675, "Test time: 14:11:00")
        c.drawString(100, 660, "Doc date: 2025-05-19")
        c.drawString(100, 645, "Doc Time: 14:11:00")
        c.drawString(100, 630, "Customer: TestCustomer")
        c.drawString(100, 615, "Item No: ITM001")
        c.drawString(100, 600, "Cust P/N: PN001")
        c.drawString(100, 585, "DESC: Sample Description")
        c.drawString(100, 570, "Qty: 5")
        c.drawString(100, 555, "COO: MALAYSIA")
        c.drawString(100, 540, "<DOCUMENT>")
        c.drawString(100, 525, "Test Report PN\tPart Number\tCustomer\tDescription\tTest report location")
        c.drawString(100, 510, "TR001\tITM001\tTestCustomer\tSample Description\tLocation1")
        c.drawString(100, 495, "</DOCUMENT>")
        c.save()
    except Exception as e:
        log_message(f"创建样本 PDF 失败: {str(e)}")
        raise
    
    globals()['PDF_SOURCE_DIR'] = test_base_dir
    return test_base_dir, sample_pdf

def log_message(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {message}\n"
    print(log_entry.strip())
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_entry)
    except PermissionError as e:
        globals()['LOG_FILE'] = os.path.join(os.path.expanduser("~"), "log.txt")
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_entry)
    except Exception as e:
        print(f"无法写入日志文件 {LOG_FILE}: {e}")

    if log_text and root:
        log_text.insert(tk.END, log_entry)
        log_text.see(tk.END)
    return log_entry.strip()

def select_source_path():
    new_path = filedialog.askdirectory(
        title="选择 PDF 源文件夹",
        initialdir=PDF_SOURCE_DIR if os.path.exists(PDF_SOURCE_DIR) else BASE_DIR
    )
    if new_path:
        globals()['PDF_SOURCE_DIR'] = new_path
        log_message(f"已锁定新的 PDF 源路径: {PDF_SOURCE_DIR}")
        path_label.config(text=f"当前路径: {PDF_SOURCE_DIR}")
        messagebox.showinfo("提示", f"已选择新的 PDF 源路径: {PDF_SOURCE_DIR}")
    else:
        log_message("用户取消了路径选择")

def initialize_data_log():
    if not os.path.exists(DATA_LOG_FILE):
        log_message(f"未找到 DataLog.xlsx: {DATA_LOG_FILE}")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "DataLog"
            headers = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            save_workbook(wb, DATA_LOG_FILE)
            log_message(f"创建新的 DataLog.xlsx 文件: {DATA_LOG_FILE}")
        except PermissionError as e:
            log_message(f"权限不足，无法创建 DataLog.xlsx: {e}")
            messagebox.showerror("错误", f"权限不足，无法创建 {DATA_LOG_FILE}")
            sys.exit(1)
    
    wb = load_workbook(DATA_LOG_FILE)
    ws = wb["DataLog"] if "DataLog" in wb else wb.active
    headers = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
    current_headers = [ws.cell(row=1, column=col).value for col in range(1, 11)]
    if not all(h in current_headers for h in headers):
        log_message(f"DataLog.xlsx 格式不正确: {DATA_LOG_FILE}")
        messagebox.showerror("错误", f"DataLog.xlsx 格式不正确，请确保包含以下表头: {', '.join(headers)}")
        sys.exit(1)
    log_message(f"DataLog.xlsx 验证通过: {DATA_LOG_FILE}")

def initialize_desc_file():
    headers = ["No.", "Customer", "Item No", "Series number", "Desc.", "Qty", "COO"]
    if not os.path.exists(DESC_FILE):
        log_message(f"未找到 desc.xlsx: {DESC_FILE}")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Description"
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            save_workbook(wb, DESC_FILE)
            log_message(f"创建新的 desc.xlsx 文件: {DESC_FILE}")
        except PermissionError as e:
            log_message(f"权限不足，无法创建 desc.xlsx: {e}")
            messagebox.showerror("错误", f"权限不足，无法创建 {DESC_FILE}")
            sys.exit(1)
    
    wb = load_workbook(DESC_FILE)
    target_sheet = next((name for name in wb.sheetnames if name.lower() == "description"), None)
    if not target_sheet:
        log_message(f"未找到 'Description' 工作表，创建新的工作表")
        ws = wb.create_sheet("Description")
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb, DESC_FILE)
    else:
        ws = wb[target_sheet]
        log_message(f"找到工作表: {target_sheet}")

    current_headers = [ws.cell(row=1, column=col).value for col in range(1, len(headers) + 1)]
    if not all(h in current_headers for h in headers):
        log_message(f"desc.xlsx 表头格式不正确: {DESC_FILE}")
        ws.delete_cols(1, ws.max_column)
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb, DESC_FILE)
        log_message(f"已更新 desc.xlsx 表头: {headers}")
    log_message(f"desc.xlsx 验证通过: {DESC_FILE}")

def test_printer(printer_name, is_priority):
    hprinter = None
    try:
        hprinter = win32print.OpenPrinter(printer_name)
        virtual_printers = ["Microsoft Print to PDF", "AnyDesk Printer", "Microsoft XPS Document Writer"]
        if any(vp in printer_name for vp in virtual_printers):
            log_message(f"跳过虚拟打印机 {printer_name} 的测试")
            return True
        if is_priority:
            test_command = ("^Q25,3\n^W50\n^H10\n^S2\nN\nT0,0,0,3,3,0,N,\"TEST PRINT\"\nP1\n")
            win32print.StartDocPrinter(hprinter, 1, ("Test Print Job", None, "RAW"))
            win32print.StartPagePrinter(hprinter)
            win32print.WritePrinter(hprinter, test_command.encode('latin-1', errors='ignore'))
            win32print.EndPagePrinter(hprinter)
            win32print.EndDocPrinter(hprinter)
            jobs = win32print.EnumJobs(hprinter, 0, 1, 2)
            if not jobs or any(job['Status'] & win32print.JOB_STATUS_COMPLETE for job in jobs):
                log_message(f"打印机 {printer_name} 测试打印成功")
                return True
            log_message(f"打印机 {printer_name} 测试打印失败")
            return False
        log_message(f"非优先打印机 {printer_name} 仅验证连接")
        return True
    except ImportError as e:
        log_message(f"模块缺失 {str(e)}, 降级 {printer_name} 为非优先")
        return False
    except Exception as e:
        log_message(f"打印机 {printer_name} 测试打印失败: {str(e)}")
        return False
    finally:
        if hprinter:
            try:
                win32print.ClosePrinter(hprinter)
            except Exception as e:
                log_message(f"关闭打印机句柄失败: {str(e)}")

def check_printer_connection(use_onenote=False):
    timeout = 5
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
            if not printers:
                default_printer = win32print.GetDefaultPrinter()
                log_message(f"未找到任何打印机，使用系统默认打印机: {default_printer}")
                if test_printer(default_printer, False):
                    return default_printer, False
                return None, False

            printer_names = [printer[2] for printer in printers]
            log_message(f"检测到的打印机列表: {printer_names}")

            priority_printers = ["Argox CX-2040 PPLB"]
            available_printers_priority = []
            available_printers_physical = []

            for printer in printers:
                printer_name = printer[2]
                if use_onenote and "onenote" in printer_name.lower() and "protected" not in printer_name.lower():
                    log_message(f"找到可用 OneNote 打印机: {printer[2]}")
                    return printer[2], False
                elif not use_onenote:
                    is_priority = any(p == printer_name for p in priority_printers)
                    test_result = test_printer(printer_name, is_priority)
                    if is_priority:
                        if test_result:
                            available_printers_priority.append((printer_name, True))
                        else:
                            log_message(f"优先打印机 {printer_name} 测试失败，但仍标记为优先")
                            available_printers_priority.append((printer_name, True))
                    elif test_result:
                        available_printers_physical.append((printer_name, False))

            if available_printers_priority:
                printer_name, is_priority = available_printers_priority[0]
                log_message(f"选择优先打印机: {printer_name}，是否优先打印机: {is_priority}")
                return printer_name, is_priority
            elif available_printers_physical:
                printer_name, is_priority = available_printers_physical[0]
                log_message(f"选择物理打印机: {printer_name}，是否优先打印机: {is_priority}")
                return printer_name, is_priority

            default_printer = win32print.GetDefaultPrinter()
            log_message(f"未找到可用打印机，使用系统默认打印机: {default_printer}")
            if test_printer(default_printer, False):
                return default_printer, False
            return None, False

        except Exception as e:
            log_message(f"检测打印机时出错: {str(e)}")
            time.sleep(1)
    log_message("检测打印机超时")
    return None, False

def clear_print_queue(printer_name):
    try:
        hprinter = win32print.OpenPrinter(printer_name, {"DesiredAccess": win32print.PRINTER_ALL_ACCESS})
        win32print.SetPrinter(hprinter, 0, None, win32print.PRINTER_CONTROL_PURGE)
        log_message(f"已清理打印机 {printer_name} 的队列")
    except Exception as e:
        log_message(f"清理打印队列失败: {str(e)}")
    finally:
        win32print.ClosePrinter(hprinter)

def check_print_job_status(printer_name, timeout=10):
    start_time = time.time()
    while time.time() - start_time < timeout:
        hprinter = win32print.OpenPrinter(printer_name)
        try:
            jobs = win32print.EnumJobs(hprinter, 0, 1, 2)
            if not jobs:
                log_message("打印任务已完成或队列为空")
                return True
            for job in jobs:
                status = job['Status']
                if status & win32print.JOB_STATUS_COMPLETE:
                    return True
                elif status & (win32print.JOB_STATUS_ERROR | win32print.JOB_STATUS_PAPEROUT):
                    log_message("打印任务失败，可能缺纸或打印机错误")
                    return False
        finally:
            win32print.ClosePrinter(hprinter)
        time.sleep(1)
    log_message("检查打印任务状态超时")
    return False

def process_pdf(pdf_path):
    if not pdf_path.lower().endswith('.pdf') or os.path.getsize(pdf_path) > 10 * 1024 * 1024:
        log_message("文件类型不支持或文件过大")
        return pd.DataFrame({"内容": ["文件类型不支持或文件过大"]})
    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_lines = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_lines.extend(text.split("\n"))
            if not all_lines:
                log_message("PDF 文件内容为空")
                return pd.DataFrame({"内容": ["PDF 文件内容为空"]})
            log_message(f"PDF 提取的总行数: {len(all_lines)}")
            return pd.DataFrame(all_lines, columns=["内容"])
    except Exception as e:
        log_message(f"处理 PDF 失败: {str(e)}")
        return pd.DataFrame({"内容": [f"PDF 处理错误: {str(e)}"]})

def initialize_template(ws):
    ws.title = "Template"
    ws["A1"] = "RADYSIS ASIA SDN.BHD."
    ws["A2"] = "Customer:"
    ws["A3"] = "Item No."
    ws["A4"] = "Cust P/N."
    ws["A5"] = "Desc."
    ws["A7"] = "Qty"
    ws["A8"] = "Test Date:"
    ws["A9"] = "Tester PIC:"
    ws["B9"] = "PROD"
    ws["B11"] = "MALAYSIA"
    ws["C9"] = "QA"
    ws["A11"] = "COO:"
    ws["C7"] = "JO NO:"
    ws["C8"] = "DC:"
    ws.merge_cells("A1:C1")
    ws.merge_cells("B2:C2")
    ws.merge_cells("B3:C3")
    ws.merge_cells("B4:C4")
    ws.merge_cells("A5:A6")
    ws.merge_cells("B5:C6")
    ws.merge_cells("A9:A10")
    ws.merge_cells("B9:B10")
    ws.merge_cells("C9:C10")
    ws.merge_cells("B11:C11")
    small_font = Font(size=8, bold=True)
    for cell in ["A1", "A2", "A3", "A4", "A5", "A7", "A8", "A9", "A11", "C7", "C8", "B9", "C9", "B11"]:
        ws[cell].font = small_font
        ws[cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 6.98
    ws.column_dimensions["B"].width = 6.98
    ws.column_dimensions["C"].width = 7.98
    for row, height in enumerate([10.78, 15.28, 11.78, 11.78, 15.28, 12.78, 15.28, 11.78, 15.28, 12.78, 10.78], 1):
        ws.row_dimensions[row].height = height
    
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in range(1, 12):
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border

    ws.print_area = "A1:C11"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.05
    ws.page_margins.right = 0.05
    ws.page_margins.top = 0.05
    ws.page_margins.bottom = 0.05
    ws.page_setup.horizontalDpi = 600
    ws.page_setup.verticalDpi = 600
    log_message("页面设置调整成功")

def save_workbook(wb, filepath):
    max_retries = 3
    for attempt in range(max_retries):
        try:
            log_message(f"尝试保存文件: {filepath} (尝试 {attempt + 1}/{max_retries})")
            with open(filepath, 'wb') as f:
                portalocker.lock(f, portalocker.LOCK_EX)
                wb.save(f)
                portalocker.unlock(f)
            log_message(f"文件保存操作完成: {filepath}")
            
            if not os.path.exists(filepath):
                raise Exception(f"文件 {filepath} 保存后未找到")
            
            wb_check = load_workbook(filepath)
            ws_check = wb_check.active
            if ws_check.max_row == 0:
                log_message(f"警告：保存的文件 {filepath} 内容为空")
                return False
            log_message(f"验证：文件 {filepath} 包含 {ws_check.max_row} 行数据")
            
            if filepath == DESC_FILE and "Description" not in wb_check.sheetnames:
                log_message(f"错误：保存后 desc.xlsx 中未找到 'Description' 工作表")
                raise Exception("保存后未找到 'Description' 工作表")
            return True
        except PermissionError as e:
            log_message(f"权限被拒绝 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(2)
                continue
            return False
        except Exception as e:
            log_message(f"保存文件失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(2)
                continue
            return False

def load_desc_data():
    try:
        wb = load_workbook(DESC_FILE, read_only=False, keep_vba=False)
        headers = ["No.", "Customer", "Item No", "Series number", "Desc.", "Qty", "COO"]
        
        all_data = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
                if row and any(cell is not None for cell in row):
                    row_dict = dict(zip(headers, row))
                    row_dict["Item No"] = str(row_dict.get("Item No", "")).strip().lower()
                    row_dict["Customer"] = str(row_dict.get("Customer", "")).strip().lower()
                    row_dict["Series number"] = str(row_dict.get("Series number", "")).strip().lower()
                    all_data.append(row_dict)
        
        df = pd.DataFrame(all_data)
        if not df.empty:
            def score_record(row):
                score = 0
                if pd.notna(row["Customer"]) and str(row["Customer"]).strip():
                    score += 3
                if pd.notna(row["COO"]) and str(row["COO"]).strip():
                    score += 2
                if pd.notna(row["Desc."]) and str(row["Desc."]).strip():
                    score += 1
                if pd.notna(row["Qty"]) and str(row["Qty"]).strip():
                    score += 1
                return score
            df["Score"] = df.apply(score_record, axis=1)
            df = df.sort_values("Score", ascending=False).drop_duplicates(subset=["Item No", "Series number"], keep="first")
        log_message(f"从 desc.xlsx 加载 {len(df)} 条唯一记录")
        return df
    except Exception as e:
        log_message(f"加载 desc.xlsx 数据失败: {str(e)}")
        return pd.DataFrame()

def find_first_lb_pdf(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                return os.path.join(root, file)
    return None

def match_desc_data(pdf_data, desc_df):
    item_no = pdf_data.get("Item No", "").strip()
    series_number = pdf_data.get("Series number", "").strip()
    customer = pdf_data.get("Customer", "").strip().lower()

    if not item_no or not series_number:
        filename = pdf_data.get("Filename", "").strip()
        match = re.match(r"(\d+-\d+-\d+)", filename)
        if match:
            item_no = match.group(1)
            pdf_data["Item No"] = item_no
        match_series = re.search(r"SN(\d+)|JO(\w+\s*\w+)", filename)
        if match_series:
            series_number = match_series.group(0) or match_series.group(1)
            pdf_data["Series number"] = series_number
        log_message(f"从文件名提取 Item No: {item_no}, Series number: {series_number}")

    if not item_no or not series_number:
        log_message("PDF 数据中未找到 Item No 或 Series number")
        return None

    item_no_cleaned = re.sub(r'[^a-zA-Z0-9-]', '', item_no).lower()
    series_number_cleaned = re.sub(r'[^a-zA-Z0-9-]', '', series_number).lower()

    if desc_df.empty:
        log_message("desc.xlsx 数据为空，无法进行匹配")
        return None

    desc_df["Item No Cleaned"] = desc_df["Item No"].astype(str).str.replace(r'[^a-zA-Z0-9-]', '', regex=True).str.lower()
    desc_df["Series number Cleaned"] = desc_df["Series number"].astype(str).str.replace(r'[^a-zA-Z0-9-]', '', regex=True).str.lower()

    matches = desc_df[(desc_df["Item No Cleaned"] == item_no_cleaned) & (desc_df["Series number Cleaned"] == series_number_cleaned)]
    if matches.empty:
        log_message(f"未在 desc.xlsx 中找到与 Item No '{item_no_cleaned}' 和 Series number '{series_number_cleaned}' 匹配的记录")
        return None

    if customer:
        matches = matches[matches["Customer"] == customer]

    coo_pdf = pdf_data.get("COO", "").strip().lower()
    matches["COO Match Score"] = matches["COO"].astype(str).str.lower().apply(lambda x: 5 if x == coo_pdf else 0)
    matches["Final Score"] = matches["Score"] + matches["COO Match Score"]

    best_match = matches.loc[matches["Final Score"].idxmax()].to_dict()
    qty_value = str(best_match.get("Qty", ""))
    if not qty_value.isdigit():
        best_match["Qty"] = "1"
    coo_value = str(best_match.get("COO", ""))
    if not coo_value.isalpha():
        best_match["COO"] = "MALAYSIA"
    log_message(f"成功匹配到 desc.xlsx 中的记录: {best_match}")
    return best_match

def backup_pdf(pdf_path, backup_dir):
    if not os.path.exists(pdf_path):
        log_message(f"源 PDF 文件不存在: {pdf_path}")
        return False

    os.makedirs(backup_dir, exist_ok=True)
    safe_filename = re.sub(r'[^a-zA-Z0-9\-\_\.]', '_', os.path.basename(pdf_path))
    backup_path = os.path.join(backup_dir, safe_filename)

    test_file = os.path.join(backup_dir, "test.txt")
    try:
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
    except Exception as e:
        log_message(f"备份目录无写入权限: {str(e)}")
        return False

    shutil.copy2(pdf_path, backup_path)
    with open(backup_path, "rb") as f:
        f.read()
    return True

def print_generic_label(printer_name, template_path, data, total_labels):
    excel = None
    wb = None
    hprinter = None
    temp_pdf_path = Path(BASE_DIR) / "temp_label.pdf"
    
    try:
        template_path = Path(template_path)
        if not template_path.exists():
            raise FileNotFoundError(f"模板文件不存在: {template_path}")

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(template_path.resolve()))
        ws = wb.ActiveSheet

        cells_to_set = {
            "B2": data.get("Customer", ""),
            "B3": data.get("Item No", ""),
            "B4": data.get("Cust P/N", ""),
            "B5": data.get("DESC", ""),
            "B7": data.get("Qty", "1"),
            "B8": data.get("Test Date", ""),
            "C8": data.get("Series number", ""),
            "C7": f"JO NO: {data.get('JQ No', '') if data.get('JQ No', '') else data.get('Lot ID', '')}",
            "B11": data.get("COO", "MALAYSIA")
        }
        small_font = Font(size=8, bold=True)
        for coord, value in cells_to_set.items():
            cell = ws[coord]
            if not isinstance(cell, MergedCell):
                cell.value = value
                cell.font = small_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.ExportAsFixedFormat(0, str(temp_pdf_path))
        log_message(f"成功导出模板为 PDF: {temp_pdf_path}")

        with open(temp_pdf_path, "rb") as f:
            pdf_data = f.read()
        log_message(f"成功读取 PDF 数据，大小: {len(pdf_data)} 字节")

        hprinter = win32print.OpenPrinter(printer_name)
        for i in range(total_labels):
            win32print.StartDocPrinter(hprinter, 1, (f"Label Print Job {i+1}", None, "RAW"))
            win32print.StartPagePrinter(hprinter)
            win32print.WritePrinter(hprinter, pdf_data)
            win32print.EndPagePrinter(hprinter)
            win32print.EndDocPrinter(hprinter)
            log_message(f"发送打印任务 {i+1}/{total_labels}")

        jobs = win32print.EnumJobs(hprinter, 0, total_labels, 2)
        if not jobs:
            log_message("打印队列为空，任务可能已完成")
        else:
            for job in jobs:
                status = job['Status']
                if status & win32print.JOB_STATUS_ERROR:
                    log_message(f"打印任务错误: Job ID {job['JobId']}")
                    raise RuntimeError("打印任务失败")

        log_message(f"通过模板打印 {total_labels} 份标签成功")

    except Exception as e:
        log_message(f"打印模板失败: {str(e)}")
        raise
    finally:
        if hprinter:
            try:
                win32print.ClosePrinter(hprinter)
                log_message(f"打印机 {printer_name} 已关闭")
            except Exception as e:
                log_message(f"关闭打印机失败: {str(e)}")
        if wb:
            try:
                wb.Close()
                log_message("Excel 工作簿已关闭")
            except Exception as e:
                log_message(f"关闭 Excel 工作簿失败: {str(e)}")
        if excel:
            try:
                excel.Quit()
                log_message("Excel 应用程序已退出")
            except Exception as e:
                log_message(f"退出 Excel 失败: {str(e)}")
        if temp_pdf_path.exists():
            try:
                temp_pdf_path.unlink()
                log_message(f"临时 PDF 文件已删除: {temp_pdf_path}")
            except Exception as e:
                log_message(f"删除临时 PDF 文件失败: {str(e)}")

def process_task():
    global auto_button, last_updated_template, root, label_count_var
    auto_button.config(state='disabled')
    try:
        if not os.path.exists(PDF_SOURCE_DIR):
            result_queue.put(("error", f"锁定 PDF 文件夹不存在: {PDF_SOURCE_DIR}"))
            return

        pdf_path = find_first_lb_pdf(PDF_SOURCE_DIR)
        if not pdf_path:
            result_queue.put(("error", f"锁定文件夹 {PDF_SOURCE_DIR} 及其子文件夹中没有 PDF 文件"))
            return

        log_message(f"选择并处理 PDF 文件: {pdf_path}")

        backup_dir = os.path.join(BASE_DIR, "backup")
        if not backup_pdf(pdf_path, backup_dir):
            result_queue.put(("error", "备份失败，程序终止"))
            return

        if not os.path.exists(last_updated_template):
            log_message(f"模板文件 {last_updated_template} 不存在，创建新模板")
            wb = Workbook()
            ws = wb.active
            initialize_template(ws)
            if not save_workbook(wb, last_updated_template):
                result_queue.put(("error", f"无法创建或保存模板文件: {last_updated_template}"))
                return
            log_message(f"创建初始模板文件成功: {last_updated_template}")

        use_onenote = use_onenote_var.get()
        printer_name, is_priority = check_printer_connection(use_onenote=use_onenote)
        if not printer_name:
            result_queue.put(("error", "未找到可用打印机！请检查打印机连接或选择物理打印机。"))
            return

        if use_onenote and "onenote" not in printer_name.lower():
            result_queue.put(("error", f"当前打印机 {printer_name} 不是 OneNote 打印机"))
            return

        # 使用 ThreadPoolExecutor 替代 timeout_decorator
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(process_pdf, pdf_path)
            try:
                df = future.result(timeout=30)  # 30 秒超时
            except concurrent.futures.TimeoutError:
                result_queue.put(("error", "PDF 处理超时"))
                return
        if df.empty or df["内容"].iloc[0].startswith("PDF"):
            result_queue.put(("error", "未在 PDF 中找到有效数据"))
            return

        target_keys = {
            "NO.": r"NO\.?\s*[:=]?\s*(\d+)",
            "Test Name": r"Test\s*Name\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
            "Series number": r"(?:S/N|Series\s*number)\s*[:=]?\s*(.*?)(?=\s*Lot\s*ID\s*[:=]|$)",
            "Lot ID": r"Lot\s*ID\s*[:=]?\s*([A-Za-z0-9-]+)",
            "JQ No": r"JQ\s*No\.?\s*[:=]?\s*([A-Za-z0-9-]+)",
            "Test Date": r"Test\s*Date\s*[:=]?\s*(\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            "Test time": r"Test\s*Time\s*[:=]?\s*(\d{1,2}:\d{2}(?::\d{2})?\s*[AP]M)?",
            "Doc date": r"Doc\s*date\s*[:=]?\s*(\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
            "Doc Time": r"Doc\s*Time\s*[:=]?\s*(\d{1,2}:\d{2}(?::\d{2})?\s*[AP]M)?",
            "Customer": r"Customer\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
            "Item No": r"Item\s*No\.?\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
            "Cust P/N": r"Cust\s*P/N\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
            "DESC": r"(?:Desc\.?|Description)\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
            "Qty": r"Qty\s*[:=]?\s*(\d+\.?\d*)",
            "COO": r"COO\s*[:=]?\s*(.*?)(?=\s{2,}|$)"
        }
        data = {}
        data["Filename"] = os.path.basename(pdf_path)

        document_data = []
        in_document_block = False
        doc_headers = ["Test Report PN", "Part Number", "Customer", "Description", "Test report location"]
        header_found = False
        for line in df["内容"]:
            line_str = str(line).strip()
            if "<DOCUMENT>" in line_str:
                in_document_block = True
                continue
            if "</DOCUMENT>" in line_str:
                in_document_block = False
                break
            if in_document_block and line_str:
                if not header_found and ("customer" in line_str.lower() or "test report pn" in line_str.lower()):
                    header_found = True
                    continue
                if header_found:
                    parts = re.split(r"\t|\s{2,}|\|", line_str)
                    if len(parts) >= len(doc_headers):
                        row_data = {doc_headers[i]: parts[i].strip() for i in range(len(doc_headers))}
                        document_data.append(row_data)

        if document_data:
            doc_row = document_data[0]
            data["Item No"] = doc_row.get("Part Number", "")
            data["Customer"] = doc_row.get("Customer", "")
            if "DESC" not in data or not data["DESC"]:
                data["DESC"] = doc_row.get("Description", "").strip()
            data["Qty"] = "1"
            data["Series number"] = ""
            data["COO"] = "MALAYSIA"

        for key, pattern in target_keys.items():
            if key in data and data[key].strip():
                continue
            for line in df["内容"]:
                match = re.search(pattern, str(line).strip(), re.IGNORECASE)
                if match:
                    value = match.group(1)
                    if key == "Test Date" and value:
                        parts = value.split('-')
                        if len(parts) > 2:
                            value = f"{parts[0]}-{parts[1]}"
                    if key == "DESC" and value:
                        data[key] = value.strip()
                        break
                    if value:
                        data[key] = value.strip()
                        break
            if key not in data:
                if key in ["Test Date", "Doc date"]:
                    data[key] = datetime.now().strftime("%Y-%m-%d")
                elif key in ["Test time", "Doc Time"]:
                    data[key] = datetime.now().strftime("%H:%M:%S")
                elif key == "Qty":
                    data[key] = "1"
                else:
                    data[key] = ""

        if not data.get("Cust P/N", "").strip():
            filename = os.path.basename(pdf_path)
            match = re.match(r"([A-Za-z0-9-]+)(?:_TestReport|_.*\.pdf)", filename)
            if match:
                data["Cust P/N"] = match.group(1)

        desc_df = load_desc_data()
        if not desc_df.empty:
            matched_data = match_desc_data(data, desc_df)
            if matched_data:
                for key in ["Customer", "DESC", "Qty", "COO"]:
                    if not data.get(key, "").strip() or data.get(key, "").lower() == "n/a":
                        desc_key = "Desc." if key == "DESC" else key
                        if desc_key in matched_data and matched_data[desc_key]:
                            data[key] = matched_data[desc_key]

        default_values = {
            "Customer": "DefaultCustomer",
            "DESC": "N/A",
            "Qty": "1",
            "COO": "MALAYSIA"
        }
        for key, default in default_values.items():
            if not data.get(key, "").strip() or data.get(key, "").lower() == "n/a":
                data[key] = default

        for date_key in ["Test Date", "Doc date"]:
            if date_key in data and data[date_key]:
                try:
                    formats = ["%Y-%m-%d", "%d/%m/%Y", "%Y%m%d"]
                    date_value = None
                    for fmt in formats:
                        try:
                            date_value = datetime.strptime(data[date_key], fmt)
                            break
                        except ValueError:
                            continue
                    data[date_key] = date_value.strftime("%Y-%m-%d") if date_value else datetime.now().strftime("%Y-%m-%d")
                except Exception:
                    data[date_key] = datetime.now().strftime("%Y-%m-%d")

        # 使用 ThreadPoolExecutor 替代 timeout_decorator
        def save_with_timeout(wb, filepath):
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(save_workbook, wb, filepath)
                try:
                    return future.result(timeout=20)
                except concurrent.futures.TimeoutError:
                    log_message(f"保存 {filepath} 超时")
                    return False

        wb_log = load_workbook(DATA_LOG_FILE)
        ws_log = wb_log["DataLog"]
        max_row_log = ws_log.max_row
        data["NO."] = max_row_log if max_row_log > 1 else 1
        new_row_log = max_row_log + 1
        headers_log = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
        for col, key in enumerate(headers_log, 1):
            ws_log.cell(row=new_row_log, column=col, value=data.get(key, "N/A"))
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for col in range(1, 11):
            ws_log.cell(row=new_row_log, column=col).border = thin_border
        if not save_with_timeout(wb_log, DATA_LOG_FILE):
            result_queue.put(("error", f"无法保存 DataLog.xlsx: {DATA_LOG_FILE}"))
            return

        wb_desc = load_workbook(DESC_FILE)
        headers_desc = ["No.", "Customer", "Item No", "Series number", "Desc.", "Qty", "COO"]
        ws_desc = wb_desc["Description"] if "Description" in wb_desc.sheetnames else wb_desc.active
        max_row_desc = ws_desc.max_row
        data["No."] = max_row_desc if max_row_desc > 1 else 1
        new_row_desc = max_row_desc + 1
        desc_data = {
            "No.": data["No."],
            "Customer": data.get("Customer", ""),
            "Item No": data.get("Item No", ""),
            "Series number": data.get("Series number", ""),
            "Desc.": data.get("DESC", ""),
            "Qty": data.get("Qty", "1"),
            "COO": data.get("COO", "MALAYSIA")
        }
        for col, key in enumerate(headers_desc, 1):
            ws_desc.cell(row=new_row_desc, column=col, value=desc_data.get(key, "N/A"))
        for col in range(1, len(headers_desc) + 1):
            ws_desc.cell(row=new_row_desc, column=col).border = thin_border
        if not save_with_timeout(wb_desc, DESC_FILE):
            result_queue.put(("error", f"无法保存 desc.xlsx: {DESC_FILE}"))
            return

        wb_updated = load_workbook(last_updated_template)
        ws_updated = wb_updated.active

        lot_id = data.get("Lot ID", "")
        jq_no = data.get("JQ No", "")
        run_number = data.get("Series number", "")
        test_date = data.get("Test Date", "")
        customer = data.get("Customer", "")
        item_no = data.get("Item No", "")
        cust_pn = data.get("Cust P/N", "")
        desc = data.get("DESC", "")
        qty = data.get("Qty", "1")
        coo = data.get("COO", "MALAYSIA")

        if cust_pn:
            cust_pn = re.sub(r'\s*Lot\s*ID\s*[:=].*$|\s*JQ\s*NO\s*[:=].*$', '', cust_pn).strip()

        cells_to_set = {
            "A1": "RADYSIS ASIA SDN.BHD.",
            "A2": "Customer:",
            "B2": customer,
            "A3": "Item No.",
            "B3": item_no,
            "A4": "Cust P/N.",
            "B4": cust_pn,
            "A5": "Desc.",
            "B5": desc,
            "A7": "Qty",
            "B7": qty,
            "A8": "Test Date:",
            "B8": test_date,
            "C8": run_number,
            "A9": "Tester PIC:",
            "B9": "PROD",
            "C9": "QA",
            "A11": "COO:",
            "B11": coo,
            "C7": f"JO NO: {jq_no if jq_no else lot_id}"
        }

        small_font = Font(size=8, bold=True)
        for coord, value in cells_to_set.items():
            cell = ws_updated[coord]
            if not isinstance(cell, MergedCell):
                cell.value = value
                cell.font = small_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws_updated.merge_cells("A1:C1")
        ws_updated.merge_cells("B2:C2")
        ws_updated.merge_cells("B3:C3")
        ws_updated.merge_cells("B4:C4")
        ws_updated.merge_cells("A5:A6")
        ws_updated.merge_cells("B5:C6")
        ws_updated.merge_cells("A9:A10")
        ws_updated.merge_cells("B9:B10")
        ws_updated.merge_cells("C9:C10")
        ws_updated.merge_cells("B11:C11")

        if not save_with_timeout(wb_updated, last_updated_template):
            result_queue.put(("error", f"无法保存模板文件: {last_updated_template}"))
            return

        wb_log_latest = load_workbook(DATA_LOG_FILE)
        ws_log_latest = wb_log_latest["DataLog"]
        latest_data = {}
        max_row_log = ws_log_latest.max_row
        if max_row_log > 1:
            headers_log = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
            for col, key in enumerate(headers_log, 1):
                latest_data[key] = ws_log_latest.cell(row=max_row_log, column=col).value or ""
        else:
            latest_data = data

        try:
            total_labels = int(label_count_var.get())
            if total_labels < 1 or total_labels > 100:
                raise ValueError("标签数量必须在 1-100 之间")
        except ValueError as e:
            total_labels = 1
            log_message(f"标签数量输入无效: {str(e)}，使用默认值: 1")

        result_queue.put(("data", (printer_name, is_priority, latest_data, pdf_path, total_labels)))

    except Exception as e:
        result_queue.put(("error", f"处理失败: {str(e)}"))
    finally:
        auto_button.config(state='normal')

def handle_input():
    try:
        # 使用 get() 替代 get_nowait()，避免队列为空时抛出异常
        result_type, message = result_queue.get(timeout=1)  # 增加超时机制
        if result_type == "data":
            printer_name, is_priority, data, pdf_path, total_labels = message
            log_message(f"准备打印: printer={printer_name}, priority={is_priority}, labels={total_labels}")
            if is_priority:
                pplb_commands = (
                    "^Q25,3\n^W50\n^H10\n^S2\nN\n"
                    f"T0,0,0,3,3,0,N,\"RADYSIS ASIA SDN.BHD.\"\n"
                    f"T0,20,0,3,3,0,N,\"Customer: {data['Customer']}\"\n"
                    f"T0,40,0,3,3,0,N,\"Item No: {data.get('Item No', '')}\"\n"
                    f"T0,60,0,3,3,0,N,\"Cust P/N: {data.get('Cust P/N', '')}\"\n"
                    f"T0,80,0,3,3,0,N,\"Desc: {data.get('DESC', '')}\"\n"
                    f"T0,100,0,3,3,0,N,\"Qty: {data.get('Qty', '1')}\"\n"
                    f"T0,120,0,3,3,0,N,\"Test Date: {data.get('Test Date', '')}    DC: {data.get('Series number', '')}\"\n"
                    f"T0,140,0,3,3,0,N,\"Tester PIC: PROD    QA\"\n"
                    f"T0,160,0,3,3,0,N,\"JO NO: {data.get('JQ No', '') if data.get('JQ No', '') else data.get('Lot ID', '')}\"\n"
                    f"T0,180,0,3,3,0,N,\"COO: {data.get('COO', 'MALAYSIA')}\"\n"
                    f"P{total_labels}\n"
                )
                hprinter = win32print.OpenPrinter(printer_name)
                try:
                    win32print.StartDocPrinter(hprinter, 1, ("Label Print Job", None, "RAW"))
                    win32print.StartPagePrinter(hprinter)
                    win32print.WritePrinter(hprinter, pplb_commands.encode('latin-1', errors='ignore'))
                    win32print.EndPagePrinter(hprinter)
                    win32print.EndDocPrinter(hprinter)
                    log_message(f"直接使用优先打印机 {printer_name} 打印 {total_labels} 份标签")
                finally:
                    win32print.ClosePrinter(hprinter)
            else:
                print_generic_label(printer_name, last_updated_template, data, total_labels)
            os.remove(pdf_path)
            log_message(f"成功删除处理后的 PDF 文件: {os.path.basename(pdf_path)}")
            result_queue.put(("success", "打印任务已完成！"))
        elif result_type in ["success", "error"]:
            messagebox.showinfo("成功", message) if result_type == "success" else messagebox.showerror("错误", message)
            auto_button.config(state='normal')
            return
    except queue.Empty:
        pass  # 队列为空，继续循环
    except Exception as e:
        log_message(f"处理队列消息失败: {str(e)}")
        result_queue.put(("error", f"处理队列消息失败: {str(e)}"))
        auto_button.config(state='normal')
    finally:
        root.after(500, handle_input)  # 增加延迟，减少 CPU 占用

def process_pdf_to_print():
    global auto_button
    auto_button.config(state='disabled')
    thread = threading.Thread(target=process_task, daemon=True)
    thread.start()

def create_gui():
    global root, log_text, auto_button, last_updated_template, use_onenote_var, path_label, label_count_var
    required_libs = {
        'pdfplumber': pdfplumber,
        'pandas': pd,
        'openpyxl': Workbook,
        'pywin32': win32print,
        'reportlab': canvas,
        'portalocker': portalocker,
        'win32com': win32com.client
    }
    missing_libs = []
    for lib, module in required_libs.items():
        try:
            if lib == 'pywin32':
                win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL)
            elif lib == 'pandas':
                if pd is None:
                    raise ImportError("pandas 未正确导入")
            elif lib == 'win32com':
                win32com.client.Dispatch("Excel.Application")
            elif module is None:
                raise ImportError(f"模块 {lib} 导入失败")
            log_message(f"{lib} 验证通过")
        except (ImportError, AttributeError, Exception) as e:
            log_message(f"依赖 {lib} 检测失败: {str(e)}")
            missing_libs.append(lib)
    if missing_libs:
        log_message(f"缺少依赖库: {', '.join(missing_libs)}")
        messagebox.showerror("错误", f"缺少以下依赖库: {', '.join(missing_libs)}\n请安装: pip install {' '.join(missing_libs)}")
        sys.exit(1)

    initialize_data_log()
    initialize_desc_file()

    test_base_dir, sample_pdf = setup_test_environment()
    log_message(f"测试环境已创建: 目录 {test_base_dir}, 样本文件 {sample_pdf}")

    if not os.path.exists(last_updated_template):
        log_message(f"模板文件 {last_updated_template} 不存在，创建新模板")
        wb = Workbook()
        ws = wb.active
        initialize_template(ws)
        if not save_workbook(wb, last_updated_template):
            log_message(f"无法创建或保存模板文件: {last_updated_template}")
            messagebox.showerror("错误", f"无法创建或保存模板文件: {last_updated_template}")
            sys.exit(1)
        log_message(f"创建初始模板文件成功: {last_updated_template}")

    root = tk.Tk()
    root.title("PDF和打印工具")
    root.geometry("400x500")
    root.protocol("WM_DELETE_WINDOW", on_closing)

    path_frame = tk.Frame(root)
    path_frame.pack(fill=tk.X, padx=5, pady=5)
    path_button = tk.Button(path_frame, text="选择 PDF 源路径", command=select_source_path)
    path_button.pack(side=tk.LEFT)
    path_label = tk.Label(path_frame, text=f"当前路径: {PDF_SOURCE_DIR}")
    path_label.pack(side=tk.LEFT, padx=5)

    use_onenote_var = tk.BooleanVar(value=False)
    use_onenote_check = tk.Checkbutton(
        root,
        text="使用 OneNote 测试",
        variable=use_onenote_var,
        onvalue=True,
        offvalue=False
    )
    use_onenote_check.pack(pady=5)

    label_frame = tk.Frame(root)
    label_frame.pack(fill=tk.X, padx=5, pady=5)
    tk.Label(label_frame, text="打印标签数量 (1-100):").pack(side=tk.LEFT)
    label_count_var = tk.StringVar(value="1")
    label_entry = tk.Entry(label_frame, textvariable=label_count_var, width=5)
    label_entry.pack(side=tk.LEFT, padx=5)

    log_frame = tk.Frame(root)
    log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    log_text = Text(log_frame, height=15, width=50)
    log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar = Scrollbar(log_frame, orient=tk.VERTICAL, command=log_text.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    log_text['yscrollcommand'] = scrollbar.set
    auto_button = tk.Button(root, text="一键完成", command=process_pdf_to_print, width=20, height=2)
    auto_button.pack(pady=5)

    root.after(100, handle_input)
    root.mainloop()

def on_closing():
    if messagebox.askokcancel("退出", "确定要退出程序吗？"):
        log_message("用户选择关闭窗口，开始清理资源")
        root.destroy()

if __name__ == "__main__":
    print("程序开始运行...")
    try:
        create_gui()
    except Exception as e:
        print(f"程序启动失败: {str(e)}")
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 程序启动失败: {str(e)}\n")