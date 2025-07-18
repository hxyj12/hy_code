import os
import shutil
import sys
import pdfplumber
import pandas as pd
import win32print
import re
import time
import threading
import queue
import portalocker
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Text, Scrollbar
from reportlab.pdfgen import canvas

# 设置 BASE_DIR 为 .exe 或脚本所在目录
BASE_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))

# 设置文件路径
LOG_FILE = os.path.join(BASE_DIR, "log.txt")
DATA_LOG_FILE = os.path.join(BASE_DIR, "data_log.xlsx")
DESC_FILE = os.path.join(BASE_DIR, "desc.xlsx")
DEFAULT_TEMPLATE_PATH = os.path.join(BASE_DIR, "updated_template.xlsx")

# 全局变量
PDF_SOURCE_DIR = r"C:\Cirris_Reports"
last_updated_template = None
log_text = None
root = None
auto_button = None
use_onenote_var = None
path_label = None
label_count_var = None

def setup_test_environment():
    test_base_dir = os.path.join(BASE_DIR, "test_Cirris_Reports")
    test_pdf_dir = os.path.join(test_base_dir, "subfolder")
    sample_pdf = os.path.join(test_pdf_dir, "LBTestSample.pdf")

    if not os.path.exists(test_base_dir):
        os.makedirs(test_base_dir)
    if not os.path.exists(test_pdf_dir):
        os.makedirs(test_pdf_dir)

    c = canvas.Canvas(sample_pdf)
    c.drawString(100, 750, "NO.: 123")
    c.drawString(100, 735, "Test Name: TestSample")
    c.drawString(100, 720, "Series number: SN12345")
    c.drawString(100, 705, "Lot ID: LOT001")
    c.drawString(100, 690, "Test Date: 2025-05-19")
    c.drawString(100, 675, "Test time: 14:11:00")
    c.drawString(100, 660, "Doc date: 2025-05-19")
    c.drawString(100, 645, "Doc Time: 14:11:00")
    c.drawString(100, 630, "Customer: TestCustomer")
    c.drawString(100, 615, "Item No: ITM001")
    c.drawString(100, 600, "Cust P/N: PN001")
    c.drawString(100, 585, "DESC: Sample Description")
    c.drawString(100, 570, "Qty: 5")
    c.drawString(100, 555, "COO: MALAYSIA")
    c.drawString(100, 540, "<DOCUMENT>")
    c.drawString(100, 525, "Test Report PN\tPart Number\tCustomer\tDescription\tTest report location")
    c.drawString(100, 510, "TR001\tITM001\tTestCustomer\tSample Description\tLocation1")
    c.drawString(100, 495, "</DOCUMENT>")
    c.save()
    
    global PDF_SOURCE_DIR
    PDF_SOURCE_DIR = test_base_dir
    return test_base_dir, sample_pdf

def log_message(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {message}\n"
    print(log_entry.strip())
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_entry)
    except Exception as e:
        print(f"无法写入日志文件 {LOG_FILE}: {e}")
    if log_text is not None:
        def update_log():
            log_text.insert(tk.END, log_entry)
            log_text.see(tk.END)
        root.after(0, update_log)
    return log_entry.strip()

def select_source_path():
    global PDF_SOURCE_DIR
    new_path = filedialog.askdirectory(
        title="选择 PDF 源文件夹",
        initialdir=PDF_SOURCE_DIR if os.path.exists(PDF_SOURCE_DIR) else BASE_DIR
    )
    if new_path:
        PDF_SOURCE_DIR = new_path
        log_message(f"已锁定新的 PDF 源路径: {PDF_SOURCE_DIR}")
        path_label.config(text=f"当前路径: {PDF_SOURCE_DIR}")
        messagebox.showinfo("提示", f"已选择新的 PDF 源路径: {PDF_SOURCE_DIR}")
    else:
        log_message("用户取消了路径选择")

def initialize_data_log():
    if not os.path.exists(DATA_LOG_FILE):
        log_message(f"未找到 DataLog.xlsx: {DATA_LOG_FILE}")
        wb = Workbook()
        ws = wb.active
        ws.title = "DataLog"
        headers = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb, DATA_LOG_FILE)
        log_message(f"创建新的 DataLog.xlsx 文件: {DATA_LOG_FILE}")
    
    wb = load_workbook(DATA_LOG_FILE)
    ws = wb["DataLog"] if "DataLog" in wb else wb.active
    headers = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
    current_headers = [ws.cell(row=1, column=col).value for col in range(1, 11)]
    if not all(h in current_headers for h in headers):
        log_message(f"DataLog.xlsx 格式不正确: {DATA_LOG_FILE}")
        messagebox.showerror("错误", f"DataLog.xlsx 格式不正确，请确保包含以下表头: {', '.join(headers)}")
        sys.exit()
    log_message(f"DataLog.xlsx 验证通过: {DATA_LOG_FILE}")

def initialize_desc_file():
    headers = ["No.", "Customer", "Item No", "Desc.", "Qty", "COO"]
    if not os.path.exists(DESC_FILE):
        log_message(f"未找到 desc.xlsx: {DESC_FILE}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Description"
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb, DESC_FILE)
        log_message(f"创建新的 desc.xlsx 文件: {DESC_FILE}")
    
    wb = load_workbook(DESC_FILE)
    target_sheet = None
    for name in wb.sheetnames:
        if name.lower() == "description":
            target_sheet = name
    if not target_sheet:
        log_message(f"未找到 'Description' 工作表，创建新的工作表")
        ws = wb.create_sheet("Description")
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb, DESC_FILE)
    else:
        ws = wb[target_sheet]
        log_message(f"找到工作表: {target_sheet}")

    current_headers = [ws.cell(row=1, column=col).value for col in range(1, len(headers) + 1)]
    if not all(h in current_headers for h in headers):
        log_message(f"desc.xlsx 表头格式不正确: {DESC_FILE}")
        ws.delete_cols(1, ws.max_column)
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        save_workbook(wb, DESC_FILE)
        log_message(f"已更新 desc.xlsx 表头: {headers}")
    log_message(f"desc.xlsx 验证通过: {DESC_FILE}")

def check_printer_connection(preferred_printers=None, use_onenote=False, require_pdf_printer=False):
    timeout = 5
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
            printer_names = [printer[2] for printer in printers]
            log_message(f"检测到的打印机列表: {printer_names}")

            if not printers:
                default_printer = win32print.GetDefaultPrinter()
                log_message(f"未找到任何打印机，使用系统默认打印机: {default_printer}")
                return default_printer

            physical_printers = ["AnyDesk Printer", "argox cx-2040", "argox cx-2140", "argox cx-3040", "argox cx-3140"]
            if not use_onenote:
                for printer in printers:
                    printer_name = printer[2].lower()
                    if any(phys in printer_name for phys in [p.lower() for p in physical_printers]):
                        hprinter = win32print.OpenPrinter(printer[2])
                        printer_info = win32print.GetPrinter(hprinter, 2)
                        win32print.ClosePrinter(hprinter)
                        if printer_info['Status'] == 0:
                            log_message(f"找到可用物理打印机: {printer[2]}")
                            return printer[2]
                        log_message(f"物理打印机 {printer[2]} 状态异常: 状态码 {printer_info['Status']}")

            if use_onenote:
                for printer in printers:
                    printer_name = printer[2].lower()
                    if "onenote" in printer_name and "protected" not in printer_name.lower():
                        hprinter = win32print.OpenPrinter(printer[2])
                        printer_info = win32print.GetPrinter(hprinter, 2)
                        win32print.ClosePrinter(hprinter)
                        if printer_info['Status'] == 0:
                            log_message(f"找到 OneNote 打印机: {printer[2]}")
                            return printer[2]
                        log_message(f"OneNote 打印机 {printer[2]} 状态异常: 状态码 {printer_info['Status']}")

            for printer in printers:
                hprinter = win32print.OpenPrinter(printer[2])
                printer_info = win32print.GetPrinter(hprinter, 2)
                win32print.ClosePrinter(hprinter)
                if printer_info['Status'] == 0:
                    log_message(f"未找到优先打印机，使用第一个可用打印机 {printer[2]}")
                    return printer[2]
                log_message(f"打印机 {printer[2]} 状态异常: 状态码 {printer_info['Status']}")

            default_printer = win32print.GetDefaultPrinter()
            log_message(f"未找到可用打印机，使用系统默认打印机: {default_printer}")
            return default_printer

        except Exception as e:
            log_message(f"检测打印机时出错: {str(e)}")
            time.sleep(1)
    log_message("检测打印机超时")
    return None

def process_pdf(pdf_path):
    if not pdf_path.lower().endswith('.pdf') or os.path.getsize(pdf_path) > 10 * 1024 * 1024:
        log_message("文件类型不支持或文件过大")
        return pd.DataFrame({"内容": ["文件类型不支持或文件过大"]})
    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_lines = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_lines.extend(text.split("\n"))
            if not all_lines:
                log_message("PDF 文件内容为空")
                return pd.DataFrame({"内容": ["PDF 文件内容为空"]})
            log_message(f"PDF 提取的总行数: {len(all_lines)}")
            df = pd.DataFrame(all_lines, columns=["内容"])
            return df
    except Exception as e:
        log_message(f"处理 PDF 失败: {str(e)}")
        return pd.DataFrame({"内容": [f"PDF 处理错误: {str(e)}"]})

def initialize_template(ws):
    ws.title = "Template"
    ws["A1"] = "RADYSIS ASIA SDN.BHD."
    ws["A2"] = "Customer:"
    ws["A3"] = "Item No."
    ws["A4"] = "Cust P/N."
    ws["A5"] = "Desc."
    ws["A7"] = "Qty"
    ws["A8"] = "Test Date:"
    ws["A9"] = "Tester PIC:"
    ws["B9"] = "PROD"
    ws["B11"] = "MALAYSIA"
    ws["C9"] = "QA"
    ws["A11"] = "COO:"
    ws["C7"] = "JO NO:"
    ws["C8"] = "DC:"
    ws.merge_cells("A1:C1")
    ws.merge_cells("B2:C2")
    ws.merge_cells("B3:C3")
    ws.merge_cells("B4:C4")
    ws.merge_cells("A5:A6")
    ws.merge_cells("B5:C6")
    ws.merge_cells("A9:A10")
    ws.merge_cells("B9:B10")
    ws.merge_cells("C9:C10")
    ws.merge_cells("B11:C11")
    small_font = Font(size=8, bold=True)
    for cell in ["A1", "A2", "A3", "A4", "A5", "A7", "A8", "A9", "A11", "C7", "C8", "B9", "C9", "B11"]:
        ws[cell].font = small_font
        ws[cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 6.98
    ws.column_dimensions["B"].width = 6.98
    ws.column_dimensions["C"].width = 7.98
    ws.row_dimensions[1].height = 10.78
    ws.row_dimensions[2].height = 15.28
    ws.row_dimensions[3].height = 11.78
    ws.row_dimensions[4].height = 11.78
    ws.row_dimensions[5].height = 15.28
    ws.row_dimensions[6].height = 12.78
    ws.row_dimensions[7].height = 15.28
    ws.row_dimensions[8].height = 11.78
    ws.row_dimensions[9].height = 15.28
    ws.row_dimensions[10].height = 12.78
    ws.row_dimensions[11].height = 10.78
    
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in range(1, 12):
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border

    ws.print_area = "A1:C11"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.05
    ws.page_margins.right = 0.05
    ws.page_margins.top = 0.05
    ws.page_margins.bottom = 0.05
    ws.page_setup.horizontalDpi = 600
    ws.page_setup.verticalDpi = 600
    log_message("页面设置调整成功")

def save_workbook(wb, filepath):
    max_retries = 2
    for attempt in range(max_retries):
        try:
            log_message(f"尝试保存文件: {filepath} (尝试 {attempt + 1}/{max_retries})")
            with open(filepath, 'wb') as f:
                portalocker.lock(f, portalocker.LOCK_EX)
                wb.save(f)
                portalocker.unlock(f)
            log_message(f"文件保存操作完成: {filepath}")
            
            if not os.path.exists(filepath):
                raise Exception(f"文件 {filepath} 保存后未找到")
            
            wb_check = load_workbook(filepath)
            ws_check = wb_check.active
            if ws_check.max_row == 0:
                log_message(f"警告：保存的文件 {filepath} 内容为空")
                return False
            log_message(f"验证：文件 {filepath} 包含 {ws_check.max_row} 行数据")
            
            if filepath == DESC_FILE:
                if "Description" not in wb_check.sheetnames:
                    log_message(f"错误：保存后 desc.xlsx 中未找到 'Description' 工作表")
                    raise Exception("保存后未找到 'Description' 工作表")
                log_message(f"验证：desc.xlsx 中 'Description' 工作表存在")
            return True
        except PermissionError as e:
            log_message(f"权限被拒绝 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(1)
                continue
            new_path = filedialog.asksaveasfilename(
                title="重新选择保存路径",
                defaultextension=".xlsx",
                initialfile=os.path.basename(filepath),
                initialdir=BASE_DIR,
                filetypes=[("Excel files", "*.xlsx")]
            )
            if new_path:
                global last_updated_template
                last_updated_template = new_path
                return save_workbook(wb, new_path)
            log_message("用户取消了重新选择保存路径")
            return False
        except Exception as e:
            log_message(f"保存文件失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(1)
                continue
            return False

def load_desc_data():
    try:
        wb = load_workbook(DESC_FILE, read_only=False, keep_vba=False)
        headers = ["No.", "Customer", "Item No", "Desc.", "Qty", "COO"]
        
        all_data = []
        for sheet_name in ["Sheet1", "Description"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
                    if row and any(cell is not None for cell in row):
                        row_dict = dict(zip(headers, row))
                        # 统一清洗 Item No 和 Customer，确保大小写和空格一致
                        row_dict["Item No"] = str(row_dict.get("Item No", "")).strip().lower()
                        row_dict["Customer"] = str(row_dict.get("Customer", "")).strip().lower()
                        all_data.append(row_dict)
        
        df = pd.DataFrame(all_data)
        if not df.empty:
            def score_record(row):
                score = 0
                if pd.notna(row["Customer"]) and str(row["Customer"]).strip():
                    score += 3
                if pd.notna(row["COO"]) and str(row["COO"]).strip():
                    score += 2
                if pd.notna(row["Desc."]) and str(row["Desc."]).strip():
                    score += 1
                if pd.notna(row["Qty"]) and str(row["Qty"]).strip():
                    score += 1
                return score
            df["Score"] = df.apply(score_record, axis=1)
            # 按 Item No 和 Customer 去重，保留最高分记录
            df = df.sort_values("Score", ascending=False).drop_duplicates(subset=["Item No", "Customer"], keep="first")
            log_message(f"desc.xlsx 去重后数据:\n{df.to_string()}")
        log_message(f"从 desc.xlsx 加载 {len(df)} 条唯一记录")
        return df
    except Exception as e:
        log_message(f"加载 desc.xlsx 数据失败: {str(e)}")
        return pd.DataFrame()

def find_first_lb_pdf(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                return os.path.join(root, file)
    return None

def match_desc_data(pdf_data, desc_df):
    item_no = pdf_data.get("Item No", "").strip()
    customer = pdf_data.get("Customer", "").strip().lower()
    if not item_no:
        log_message("PDF 数据中未找到 Item No")
        return None

    # 统一清洗 Item No
    item_no_cleaned = re.sub(r'[^a-zA-Z0-9-]', '', item_no).lower()
    log_message(f"PDF Item No 清洗后: {item_no_cleaned}")
    if desc_df.empty:
        log_message("desc.xlsx 数据为空，无法进行匹配")
        return None

    desc_df["Item No Cleaned"] = desc_df["Item No"].astype(str).str.replace(r'[^a-zA-Z0-9-]', '', regex=True).str.lower()
    log_message(f"desc.xlsx Item No 清洗后:\n{desc_df[['Item No', 'Item No Cleaned']].to_string()}")

    # 优先按 Item No 匹配，忽略空 Customer
    matches = desc_df[desc_df["Item No Cleaned"] == item_no_cleaned]
    if matches.empty:
        log_message(f"未在 desc.xlsx 中找到与 Item No '{item_no_cleaned}' 匹配的记录")
        return None

    # 如果有 Customer，过滤匹配的记录
    if customer:
        matches_with_customer = matches[matches["Customer"] == customer]
        if not matches_with_customer.empty:
            matches = matches_with_customer
            log_message(f"Customer 匹配成功，筛选后记录:\n{matches.to_string()}")

    coo_pdf = pdf_data.get("COO", "").strip().lower()
    matches["COO Match Score"] = matches["COO"].astype(str).str.lower().apply(lambda x: 5 if x == coo_pdf else 0)
    matches["Final Score"] = matches["Score"] + matches["COO Match Score"]

    best_match = matches.loc[matches["Final Score"].idxmax()].to_dict()
    log_message(f"成功匹配到 desc.xlsx 中的记录: {best_match}")
    return best_match

def backup_pdf(pdf_path, backup_dir):
    if not os.path.exists(pdf_path):
        log_message(f"源 PDF 文件不存在: {pdf_path}")
        return False

    log_message(f"源 PDF 文件存在，大小: {os.path.getsize(pdf_path)} 字节")
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir, exist_ok=True)
        log_message(f"创建备份目录: {backup_dir}")

    safe_filename = re.sub(r'[^a-zA-Z0-9\-\_\.]', '_', os.path.basename(pdf_path))
    backup_path = os.path.join(backup_dir, safe_filename)
    log_message(f"目标备份路径: {backup_path}")

    test_file = os.path.join(backup_dir, "test.txt")
    try:
        with open(test_file, "w") as f:
            f.write("test")
        os.remove(test_file)
    except Exception as e:
        log_message(f"备份目录无写入权限: {str(e)}")
        return False

    shutil.copy2(pdf_path, backup_path)
    log_message(f"通过文件复制备份: 从 {pdf_path} 到 {backup_path}")
    backup_size = os.path.getsize(backup_path)
    log_message(f"验证：备份文件存在，大小: {backup_size} 字节")

    with open(backup_path, "rb") as f:
        f.read()
    log_message("备份文件完整性验证通过")

    os.rename(backup_path, backup_path + ".tmp")
    os.rename(backup_path + ".tmp", backup_path)
    log_message("备份文件读写验证通过")
    return True

def process_pdf_to_print():
    global auto_button, last_updated_template, PDF_SOURCE_DIR, root, label_count_var
    auto_button.config(state='disabled')
    result_queue = queue.Queue()

    def process_task():
        hprinter = None
        try:
            if not os.path.exists(PDF_SOURCE_DIR):
                result_queue.put(("error", f"锁定 PDF 文件夹不存在: {PDF_SOURCE_DIR}"))
                return

            pdf_path = find_first_lb_pdf(PDF_SOURCE_DIR)
            if not pdf_path:
                result_queue.put(("error", f"锁定文件夹 {PDF_SOURCE_DIR} 及其子文件夹中没有 PDF 文件"))
                return

            log_message(f"选择并处理 PDF 文件: {pdf_path}")

            backup_dir = os.path.join(BASE_DIR, "backup")
            log_message(f"备份目录路径: {backup_dir}")
            if not backup_pdf(pdf_path, backup_dir):
                result_queue.put(("error", "备份失败，程序终止"))
                return

            global last_updated_template
            if last_updated_template is None or not os.path.exists(last_updated_template):
                last_updated_template = DEFAULT_TEMPLATE_PATH
                if not os.path.exists(last_updated_template):
                    wb = Workbook()
                    ws = wb.active
                    initialize_template(ws)
                    save_workbook(wb, last_updated_template)
                    log_message(f"创建初始模板文件成功: {last_updated_template}")

            use_onenote = use_onenote_var.get()
            printer_name = check_printer_connection(use_onenote=use_onenote)
            if not printer_name:
                result_queue.put(("error", "未找到可用打印机！"))
                return

            if use_onenote and not any(onename in printer_name.lower() for onename in ["onenote"]):
                result_queue.put(("error", f"当前打印机 {printer_name} 不是 OneNote 打印机"))
                return

            log_message(f"开始提取 {pdf_path} 数据...")
            df = process_pdf(pdf_path)
            if df.empty or df["内容"].iloc[0].startswith("PDF"):
                result_queue.put(("error", "未在 PDF 中找到有效数据"))
                return

            log_message(f"PDF 提取的完整内容:\n{df['内容'].to_string()}")

            target_keys = {
                "NO.": r"NO\.?\s*[:=]?\s*(\d+)|No\.?\s*[:=]?\s*(\d+)|Number\s*[:=]?\s*(\d+)",
                "Test Name": r"Test\s*Name\s*[:=]?\s*(.*?)(?=\s{2,}|$)|TestName\s[:=]?\s*(.*?)(?=\s{2,}|$)",
                "Series number": r"S/N\s*[:=]?\s*(.*?)(?=\s*Lot\s*ID\s*[:=]|$)|SN\s*[:=]?\s*(.*?)(?=\s*Lot\s*ID\s*[:=]|$)|Serial\s*Number\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Run\s*Number\s*[:=]?\s*(\d+)",
                "Lot ID": r"Lot\s*ID\s*[:=]?\s*([A-Za-z0-9-]+)|lot\s*id\s*[:=]?\s*([A-Za-z0-9-]+)|LOT\s*ID\s*[:=]?\s*([A-Za-z0-9-]+)|Lot\s*[:=]?\s*([A-Za-z0-9-]+)|LOTID\s*[:=]?\s*([A-Za-z0-9-]+)|Lot-ID\s*[:=]?\s*([A-Za-z0-9-]+)",
                "JQ No": r"JQ\s*No\.?\s*[:=]?\s*([A-Za-z0-9-]+)",
                "Test Date": r"Test\s*Date\s*[:=]?\s*(\d{4}-\d{4}(?=(?:-\d+)?(?:\s|$))|\d{4}-\d{2}-\d{2}\s*\|\s*\d{1,2}:\d{2}:\d{2}\s*[AP]M|\d{1,2}/\d{1,2}/\d{4}\s*\d{1,2}:\d{2}:\d{2}\s*[AP]M|\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{2}-\d{2}-\d{4})",
                "Test time": r"Test\s*Time\s*[:=]?\s*(\d{1,2}:\d{2}(?::\d{2})?\s*[AP]M)|Time\s*of\s*Test\s*[:=]?\s*(\d{1,2}:\d{2}(?::\d{2})?\s*[AP]M)|\d{1,2}:\d{2}:\d{2}\s*[AP]M",
                "Doc date": r"Doc\s*date\s*[:=]?\s*(\d{4}-\d{2}-\d{2}\s*\|\s*\d{1,2}:\d{2}:\d{2}\s*[AP]M|\d{1,2}/\d{1,2}/\d{4}\s*\d{1,2}:\d{2}:\d{2}\s*[AP]M|\d{4}-\d{2}-\d{2}|\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
                "Doc Time": r"Doc\s*Time\s*[:=]?\s*(\d{1,2}:\d{2}(?::\d{2})?\s*[AP]M)|Document\s*Time\s*[:=]?\s*(\d{1,2}:\d{2}(?::\d{2})?\s*[AP]M)|\d{1,2}:\d{2}:\d{2}\s*[AP]M",
                "Customer": r"Customer\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Cust\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Customer\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
                "Item No": r"Item\s*No\.?\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Item\s*Number\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Test\s*Name\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
                "Cust P/N": r"Cust\s*P/N\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Customer\s*Part\s*Number\s*[:=]?\s*(.*?)(?=\s{2,}|$)|P/N\s*[:=]?\s*(.*?)(?=\s{2,}|$)",
                "DESC": r"(?:Desc\.?|Description|Comment)\s*[:=：;-]?\s*(.*?)(?=\s{2,}|\n|$|\Z)|(?:Desc\.?|Description|Comment)\s*(.*?)(?=\s{2,}|\n|$|\Z)",
                "Qty": r"Qty\s*[:=]?\s*(\d+\.?\d*)|Quantity\s*[:=]?\s*(\d+\.?\d*)|Cable\s*Number\s*[:=]?\s*(\d+)|QTY\s*[:=]?\s*(\d+\.?\d*)",
                "COO": r"COO\s*[:=]?\s*(.*?)(?=\s{2,}|$)|Country\s*of\s*Origin\s*[:=]?\s*(.*?)(?=\s{2,}|$)"
            }
            data = {}

            document_data = []
            in_document_block = False
            doc_headers = ["Test Report PN", "Part Number", "Customer", "Description", "Test report location"]
            header_found = False
            for line in df["内容"]:
                line_str = str(line).strip()
                if "<DOCUMENT>" in line_str:
                    in_document_block = True
                    continue
                if "</DOCUMENT>" in line_str:
                    in_document_block = False
                    break
                if in_document_block and line_str:
                    if not header_found and ("customer" in line_str.lower() or "test report pn" in line_str.lower()):
                        header_found = True
                        log_message(f"找到 <DOCUMENT> 表头: {line_str}")
                        continue
                    if header_found:
                        parts = re.split(r"\t|\s{2,}|\|", line_str)
                        if len(parts) >= len(doc_headers):
                            row_data = {doc_headers[i]: parts[i].strip() for i in range(len(doc_headers))}
                            document_data.append(row_data)
                            log_message(f"提取 <DOCUMENT> 行数据: {row_data}")

            if document_data:
                doc_row = document_data[0]
                data["Item No"] = doc_row.get("Part Number", "")
                data["Customer"] = doc_row.get("Customer", "")
                if "DESC" not in data or not data["DESC"]:
                    desc_from_doc = doc_row.get("Description", "").strip()
                    if desc_from_doc:
                        data["DESC"] = desc_from_doc
                        log_message(f"从 <DOCUMENT> 提取 DESC: {desc_from_doc}")
                data["Qty"] = "1"
                data["Series number"] = ""
                data["COO"] = "MALAYSIA"

            for key, pattern in target_keys.items():
                if key in data and data[key].strip():
                    continue
                for line in df["内容"]:
                    line_str = str(line).strip()
                    match = re.search(pattern, line_str, re.IGNORECASE)
                    if match:
                        value = next((group for group in match.groups() if group is not None), None)
                        if key == "Test Date" and value:
                            parts = value.split('-')
                            if len(parts) > 2:
                                value = f"{parts[0]}-{parts[1]}"
                                log_message(f"修正 Test Date: {value}")
                        if key == "DESC" and value is not None:
                            if value.strip():
                                data[key] = value.strip()
                            break
                        if value:
                            data[key] = value.strip()
                            break
                if key not in data:
                    if key in ["Test Date", "Doc date"]:
                        data[key] = datetime.now().strftime("%Y-%m-%d")
                    elif key in ["Test time", "Doc Time"]:
                        data[key] = datetime.now().strftime("%H:%M:%S")
                    elif key == "Qty":
                        data[key] = "1"
                    else:
                        data[key] = ""

            log_message(f"PDF 提取后的 data 字典: {data}")

            desc_df = load_desc_data()
            if not desc_df.empty:
                matched_data = match_desc_data(data, desc_df)
                if matched_data:
                    for key in ["Customer", "DESC", "Qty", "COO"]:
                        if not data.get(key, "").strip() or data.get(key, "").lower() == "n/a":
                            desc_key = "Desc." if key == "DESC" else key
                            if desc_key in matched_data and matched_data[desc_key]:
                                data[key] = matched_data[desc_key]
                                log_message(f"从 desc.xlsx 补全 {key}: {data[key]}")
                else:
                    log_message("未找到匹配的 desc.xlsx 数据")
            else:
                log_message("desc.xlsx 数据为空，无法进行匹配")

            default_values = {
                "Customer": "DefaultCustomer",
                "DESC": "N/A",
                "Qty": "1",
                "COO": "MALAYSIA"
            }
            for key, default in default_values.items():
                if not data.get(key, "").strip() or data.get(key, "").lower() == "n/a":
                    data[key] = default
                    log_message(f"未找到 {key} 的值，设置为默认值: {default}")

            for date_key in ["Test Date", "Doc date"]:
                if date_key in data and data[date_key]:
                    date_str = data[date_key]
                    try:
                        if date_key == "Test Date" and re.match(r"^\d{4}-\d{4}$", date_str):
                            continue
                        formats = [
                            "%Y-%m-%d | %I:%M:%S %p",
                            "%m/%d/%Y %I:%M:%S %p",
                            "%Y-%m-%d",
                            "%d/%m/%Y",
                            "%m/%d/%Y",
                            "%Y/%m/%d",
                            "%d-%m-%Y",
                            "%Y%m%d"
                        ]
                        date_value = None
                        for fmt in formats:
                            try:
                                date_value = datetime.strptime(date_str, fmt)
                                break
                            except ValueError:
                                continue
                        if date_value:
                            data[date_key] = date_value.strftime("%Y-%m-%d")
                        else:
                            data[date_key] = datetime.now().strftime("%Y-%m-%d")
                    except Exception:
                        data[date_key] = datetime.now().strftime("%Y-%m-%d")

            wb_log = load_workbook(DATA_LOG_FILE)
            ws_log = wb_log["DataLog"]
            max_row_log = ws_log.max_row
            data["NO."] = max_row_log if max_row_log > 1 else 1
            new_row_log = max_row_log + 1
            headers_log = ["NO.", "Customer", "Test Name", "Series number", "Lot ID", "Test Date", "Test time", "DESC", "Doc date", "Doc Time"]
            for col, key in enumerate(headers_log, 1):
                value = data.get(key, "N/A")
                ws_log.cell(row=new_row_log, column=col, value=value)
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            for col in range(1, 11):
                ws_log.cell(row=new_row_log, column=col).border = thin_border
            save_workbook(wb_log, DATA_LOG_FILE)

            wb_desc = load_workbook(DESC_FILE)
            headers_desc = ["No.", "Customer", "Item No", "Desc.", "Qty", "COO"]
            ws_desc = wb_desc["Description"] if "Description" in wb_desc.sheetnames else wb_desc.active
            save_workbook(wb_desc, DESC_FILE)

            max_row_desc = ws_desc.max_row
            data["No."] = max_row_desc if max_row_desc > 1 else 1
            new_row_desc = max_row_desc + 1
            desc_data = {
                "No.": data["No."],
                "Customer": data.get("Customer", ""),
                "Item No": data.get("Item No", ""),
                "Desc.": data.get("DESC", ""),
                "Qty": data.get("Qty", "1"),
                "COO": data.get("COO", "MALAYSIA")
            }
            for col, key in enumerate(headers_desc, 1):
                value = desc_data.get(key, "N/A")
                ws_desc.cell(row=new_row_desc, column=col, value=value)
            for col in range(1, len(headers_desc) + 1):
                ws_desc.cell(row=new_row_desc, column=col).border = thin_border
            save_workbook(wb_desc, DESC_FILE)

            wb_updated = load_workbook(last_updated_template)
            ws_updated = wb_updated.active

            lot_id = data.get("Lot ID", "")
            jq_no = data.get("JQ No", "")
            run_number = data.get("Series number", "")
            test_date = data.get("Test Date", "")
            customer = data.get("Customer", "")
            item_no = data.get("Item No", "")
            cust_pn = data.get("Cust P/N", "")
            desc = data.get("DESC", "")
            qty = data.get("Qty", "1")
            coo = data.get("COO", "MALAYSIA")

            if cust_pn:
                cust_pn = re.sub(r'\s*Lot\s*ID\s*[:=].*$|\s*JQ\s*NO\s*[:=].*$', '', cust_pn).strip()

            cells_to_set = {
                "A1": "RADYSIS ASIA SDN.BHD.",
                "A2": "Customer:",
                "B2": customer,
                "A3": "Item No.",
                "B3": item_no,
                "A4": "Cust P/N.",
                "B4": cust_pn,
                "A5": "Desc.",
                "B5": desc,
                "A7": "Qty",
                "B7": qty,
                "A8": "Test Date:",
                "B8": test_date,
                "C8": run_number,
                "A9": "Tester PIC:",
                "B9": "PROD",
                "C9": "QA",
                "A11": "COO:",
                "B11": coo,
                "C7": f"JO NO: {jq_no if jq_no else lot_id}"
            }

            small_font = Font(size=8, bold=True)
            for coord, value in cells_to_set.items():
                cell = ws_updated[coord]
                if not isinstance(cell, MergedCell):
                    cell.value = value
                    cell.font = small_font
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            ws_updated.merge_cells("A1:C1")
            ws_updated.merge_cells("B2:C2")
            ws_updated.merge_cells("B3:C3")
            ws_updated.merge_cells("B4:C4")
            ws_updated.merge_cells("A5:A6")
            ws_updated.merge_cells("B5:C6")
            ws_updated.merge_cells("A9:A10")
            ws_updated.merge_cells("B9:B10")
            ws_updated.merge_cells("C9:C10")
            ws_updated.merge_cells("B11:C11")

            save_workbook(wb_updated, last_updated_template)

            try:
                total_labels = int(label_count_var.get())
                if total_labels < 1 or total_labels > 100:
                    raise ValueError("标签数量必须在 1-100 之间")
            except ValueError:
                total_labels = 1
                log_message("标签数量输入无效，使用默认值: 1")

            result_queue.put(("data", (printer_name, data, pdf_path, total_labels)))

        except Exception as e:
            result_queue.put(("error", f"处理失败: {str(e)}"))
        finally:
            if hprinter:
                try:
                    win32print.ClosePrinter(hprinter)
                except Exception as e:
                    log_message(f"关闭打印机句柄失败: {str(e)}")
            root.after(0, handle_input)

    def handle_input():
        try:
            result_type, message = result_queue.get(timeout=20)
            if result_type == "data":
                printer_name, data, pdf_path, total_labels = message
                print_content = (
                    f"RADYSIS ASIA SDN.BHD.\n"
                    f"Customer: {data['Customer']}\n"
                    f"Item No. {data['Item No']}\n"
                    f"Cust P/N. {data.get('Cust P/N', '')}\n"
                    f"Desc. {data['DESC']}\n"
                    f"Qty {data['Qty']}\n"
                    f"Test Date: {data['Test Date']}    DC: {data.get('Series number', '')}\n"
                    f"Tester PIC: PROD    QA\n"
                    f"JO NO: {data.get('JQ No', '') if data.get('JQ No', '') else data.get('Lot ID', '')}\n"
                    f"COO: {data['COO']}\n"
                )

                max_retries = 2
                hprinter = None
                for attempt in range(max_retries):
                    try:
                        if not os.path.exists(last_updated_template):
                            raise FileNotFoundError(f"模板文件不存在: {last_updated_template}")

                        hprinter = win32print.OpenPrinter(printer_name)
                        printer_info = win32print.GetPrinter(hprinter, 2)
                        if printer_info['Status'] != 0:
                            raise RuntimeError(f"打印机 {printer_name} 状态异常，状态码: {printer_info['Status']}")

                        win32print.SetDefaultPrinter(printer_name)
                        win32print.StartDocPrinter(hprinter, 1, ("Label Print Job", None, "RAW"))
                        win32print.StartPagePrinter(hprinter)

                        for i in range(total_labels):
                            start_time = time.time()
                            win32print.WritePrinter(hprinter, print_content.encode('utf-8'))
                            log_message(f"打印第 {i+1}/{total_labels} 页")
                            if time.time() - start_time > 5:
                                raise TimeoutError(f"打印第 {i+1} 页超时")

                        win32print.EndPagePrinter(hprinter)
                        win32print.EndDocPrinter(hprinter)
                        log_message("打印完成")

                        start_time = time.time()
                        os.remove(pdf_path)
                        if time.time() - start_time > 5:
                            raise TimeoutError(f"删除文件 {pdf_path} 超时")
                        log_message(f"成功删除处理后的 PDF 文件: {os.path.basename(pdf_path)}")
                        result_queue.put(("success", f"已成功处理并删除 PDF 文件: {os.path.basename(pdf_path)}"))
                        break

                    except FileNotFoundError as fnf_err:
                        log_message(f"打印失败 (尝试 {attempt + 1}/{max_retries}): {str(fnf_err)}")
                        if attempt < max_retries - 1:
                            time.sleep(1)
                            continue
                        result_queue.put(("error", f"打印失败: {str(fnf_err)}"))
                    except TimeoutError as te:
                        log_message(f"操作超时 (尝试 {attempt + 1}/{max_retries}): {str(te)}")
                        if attempt < max_retries - 1:
                            time.sleep(1)
                            continue
                        result_queue.put(("error", f"操作超时: {str(te)}"))
                    except Exception as e:
                        log_message(f"打印失败 (尝试 {attempt + 1}/{max_retries}): {str(e)}")
                        if attempt < max_retries - 1:
                            time.sleep(1)
                            continue
                        result_queue.put(("error", f"打印失败: {str(e)}"))
                    finally:
                        if hprinter:
                            try:
                                win32print.ClosePrinter(hprinter)
                                hprinter = None
                            except Exception as e:
                                log_message(f"关闭打印机句柄失败: {str(e)}")

                result_queue.put(("success", "打印任务已完成！"))

            elif result_type in ["success", "error"]:
                messagebox.showinfo("成功", message) if result_type == "success" else \
                messagebox.showerror("错误", message)
                auto_button.config(state='normal')
                return

            root.after(100, handle_input)

        except queue.Empty:
            log_message("队列处理超时")
            result_queue.put(("error", "队列处理超时"))
            auto_button.config(state='normal')
        except Exception as e:
            log_message(f"处理队列消息失败: {str(e)}")
            result_queue.put(("error", f"处理队列消息失败: {str(e)}"))
            auto_button.config(state='normal')

    thread = threading.Thread(target=process_task)
    thread.daemon = True
    thread.start()

def create_gui():
    global root, log_text, auto_button, last_updated_template, use_onenote_var, path_label, label_count_var
    initialize_data_log()
    initialize_desc_file()

    test_base_dir, sample_pdf = setup_test_environment()
    log_message(f"测试环境已创建: 目录 {test_base_dir}, 样本文件 {sample_pdf}")

    global last_updated_template
    last_updated_template = DEFAULT_TEMPLATE_PATH
    if not os.path.exists(last_updated_template):
        wb = Workbook()
        ws = wb.active
        initialize_template(ws)
        save_workbook(wb, last_updated_template)
        log_message(f"程序启动时创建初始模板文件成功: {last_updated_template}")
    else:
        log_message(f"程序启动时检测到现有模板文件: {last_updated_template}")

    root = tk.Tk()
    root.title("PDF和打印工具")
    root.geometry("400x500")
    root.protocol("WM_DELETE_WINDOW", on_closing)

    path_frame = tk.Frame(root)
    path_frame.pack(fill=tk.X, padx=5, pady=5)
    path_button = tk.Button(path_frame, text="选择 PDF 源路径", command=select_source_path)
    path_button.pack(side=tk.LEFT)
    path_label = tk.Label(path_frame, text=f"当前路径: {PDF_SOURCE_DIR}")
    path_label.pack(side=tk.LEFT, padx=5)

    use_onenote_var = tk.BooleanVar()
    use_onenote_check = tk.Checkbutton(
        root,
        text="使用 OneNote 测试",
        variable=use_onenote_var,
        onvalue=True,
        offvalue=False
    )
    use_onenote_check.pack(pady=5)

    label_frame = tk.Frame(root)
    label_frame.pack(fill=tk.X, padx=5, pady=5)
    tk.Label(label_frame, text="打印标签数量 (1-100):").pack(side=tk.LEFT)
    label_count_var = tk.StringVar(value="1")
    label_entry = tk.Entry(label_frame, textvariable=label_count_var, width=5)
    label_entry.pack(side=tk.LEFT, padx=5)

    log_frame = tk.Frame(root)
    log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    log_text = Text(log_frame, height=15, width=50)
    log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar = Scrollbar(log_frame, orient=tk.VERTICAL, command=log_text.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    log_text['yscrollcommand'] = scrollbar.set
    auto_button = tk.Button(root, text="一键完成", command=process_pdf_to_print, width=20, height=2)
    auto_button.pack(pady=5)
    root.mainloop()

def on_closing():
    if messagebox.askokcancel("退出", "确定要退出程序吗？"):
        log_message("用户选择关闭窗口，开始清理资源")
        for thread in threading.enumerate():
            if thread != threading.current_thread():
                log_message(f"发现存活线程: {thread.name}")
                thread.daemon = True

        printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
        for printer in printers:
            try:
                hprinter = win32print.OpenPrinter(printer[2])
                win32print.ClosePrinter(hprinter)
                log_message(f"关闭打印机句柄: {printer[2]}")
            except Exception as e:
                log_message(f"关闭打印机 {printer[2]} 句柄失败: {str(e)}")

        log_message("资源清理完成，程序退出")
        root.destroy()

if __name__ == "__main__":
    print("程序开始运行...")
    try:
        create_gui()
    except Exception as e:
        print(f"程序启动失败: {str(e)}")
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 程序启动失败: {str(e)}\n")